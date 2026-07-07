# site_audit/report.py
"""
Генерация отчётов: Excel (openpyxl) и HTML.

Excel-файл содержит:
  - Лист «Сводка» с общей статистикой
  - Отдельный лист для каждой проверки
  - Автоширина колонок, фильтры, закреплённая шапка
"""

from __future__ import annotations

import html as html_mod
import json
import os
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Названия листов ─────────────────────────────────────────────────────────

_SHEET_NAMES = {
    "empty_pages": "Пустые страницы",
    "seo": "SEO",
    "broken_links": "Битые ссылки",
    "images": "Картинки",
    "redirects": "Редиректы",
    "duplicates": "Дубликаты",
    "placeholders": "Заглушки",
}

# Приоритетные колонки (выводятся первыми)
_PRIORITY_COLUMNS = ["check", "url", "page_url", "source_url", "target_url", "src"]

# Колонки, которые не нужны в Excel
_SKIP_COLUMNS = {"check"}

# Стили
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
_CELL_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
_THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
_SUMMARY_HEADER_FONT = Font(bold=True, size=14, color="2F5496")
_SUMMARY_LABEL_FONT = Font(bold=True, size=11)
_GOOD_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
_BAD_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")


# ═════════════════════════════════════════════════════════════════════════════
# Excel
# ═════════════════════════════════════════════════════════════════════════════

def save_excel(results: dict[str, list[dict]],
               summaries: dict[str, str],
               *,
               base_url: str = "",
               filepath: str = "audit_report.xlsx"):
    """Сохраняет отчёт в Excel с отдельным листом для каждой проверки."""
    wb = Workbook()

    # ── Лист «Сводка» ──────────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Сводка"
    _build_summary_sheet(ws_summary, results, base_url)

    # ── Листы для каждой проверки ───────────────────────────────────────
    for check_name, rows in results.items():
        sheet_title = _SHEET_NAMES.get(check_name, check_name)[:31]  # Excel лимит 31 символ
        ws = wb.create_sheet(title=sheet_title)

        if not rows:
            ws.append(["Проблем не найдено"])
            ws["A1"].font = Font(italic=True, color="808080")
            continue

        _build_data_sheet(ws, rows, check_name)

    # ── Сохраняем ──────────────────────────────────────────────────────
    wb.save(filepath)
    print(f"  [report] Excel сохранён: {filepath}")


def _build_summary_sheet(ws, results: dict[str, list[dict]], base_url: str):
    """Заполняет лист «Сводка»."""
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    total_issues = sum(len(rows) for rows in results.values())

    # Заголовок
    ws.merge_cells("A1:D1")
    ws["A1"] = "Аудит сайта"
    ws["A1"].font = _SUMMARY_HEADER_FONT

    ws["A3"] = "Сайт:"
    ws["A3"].font = _SUMMARY_LABEL_FONT
    ws["B3"] = base_url

    ws["A4"] = "Дата:"
    ws["A4"].font = _SUMMARY_LABEL_FONT
    ws["B4"] = now

    ws["A5"] = "Всего проблем:"
    ws["A5"].font = _SUMMARY_LABEL_FONT
    ws["B5"] = total_issues

    # Таблица по проверкам
    ws["A7"] = "Проверка"
    ws["B7"] = "Проблем"
    ws["C7"] = "Статус"
    for col in range(1, 4):
        cell = ws.cell(row=7, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT

    row_num = 8
    for check_name, rows in results.items():
        label = _SHEET_NAMES.get(check_name, check_name)
        count = len(rows)

        ws.cell(row=row_num, column=1, value=label)
        ws.cell(row=row_num, column=2, value=count)

        status_cell = ws.cell(row=row_num, column=3)
        if count == 0:
            status_cell.value = "✓ OK"
            status_cell.fill = _GOOD_FILL
        else:
            status_cell.value = f"✗ {count} проблем"
            status_cell.fill = _BAD_FILL

        row_num += 1

    # Ширина колонок
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 20


def _build_data_sheet(ws, rows: list[dict], check_name: str):
    """Заполняет лист данными одной проверки."""
    # Определяем колонки
    columns = _get_ordered_columns(rows)

    # Человекочитаемые названия
    col_labels = [_humanize_column(c) for c in columns]

    # Шапка
    for col_idx, label in enumerate(col_labels, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT
        cell.border = _THIN_BORDER

    # Данные
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, col_name in enumerate(columns, 1):
            value = row_data.get(col_name, "")
            cell_value = _format_value(value)
            cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
            cell.alignment = _CELL_ALIGNMENT
            cell.border = _THIN_BORDER

    # Автоширина
    _auto_width(ws, columns, rows)

    # Фильтры
    ws.auto_filter.ref = ws.dimensions

    # Закрепляем шапку
    ws.freeze_panes = "A2"


def _get_ordered_columns(rows: list[dict]) -> list[str]:
    """Собирает колонки в нужном порядке: приоритетные первыми."""
    seen: set[str] = set()
    ordered: list[str] = []

    # Приоритетные
    for col in _PRIORITY_COLUMNS:
        if any(col in r for r in rows) and col not in _SKIP_COLUMNS:
            ordered.append(col)
            seen.add(col)

    # Остальные в порядке появления
    for row in rows:
        for col in row:
            if col not in seen and col not in _SKIP_COLUMNS:
                ordered.append(col)
                seen.add(col)

    return ordered


def _format_value(value) -> str | int | float | None:
    """Конвертирует значение для ячейки Excel."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "да" if value else ""
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, list):
        if not value:
            return ""
        # Список строк — через перенос строки
        parts = []
        for item in value:
            if isinstance(item, dict):
                parts.append(json.dumps(item, ensure_ascii=False, default=str))
            else:
                parts.append(str(item))
        return "\n".join(parts)
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False, default=str)
    return str(value)


_COLUMN_LABELS = {
    "url": "URL",
    "page_url": "Страница",
    "source_url": "Источник ссылки",
    "target_url": "Целевой URL",
    "src": "URL картинки",
    "status_code": "HTTP-статус",
    "final_url": "Финальный URL",
    "html_size": "Размер HTML (байт)",
    "text_length": "Длина текста (симв.)",
    "is_empty": "Пустая?",
    "reason": "Причина",
    "issues": "Проблемы",
    "issues_count": "Кол-во проблем",
    "title": "Title",
    "title_length": "Длина Title",
    "meta_description": "Meta Description",
    "meta_description_length": "Длина Description",
    "canonical": "Canonical",
    "canonical_mismatch": "Canonical ≠ URL?",
    "meta_robots": "Meta Robots",
    "has_noindex": "Noindex?",
    "has_nofollow": "Nofollow?",
    "h1_count": "Кол-во H1",
    "h1_text": "Текст H1",
    "og": "Open Graph",
    "jsonld_count": "JSON-LD (кол-во)",
    "jsonld_types": "JSON-LD типы",
    "link_type": "Тип ссылки",
    "error": "Ошибка",
    "content_length": "Размер файла (байт)",
    "chain_length": "Хопов",
    "chain": "Цепочка",
    "dup_type": "Тип дубля",
    "value": "Значение",
    "urls": "URL-дубликаты",
    "count": "Кол-во дублей",
    "findings": "Находки",
    "findings_count": "Кол-во находок",
    "severity": "Важность",
}


def _humanize_column(col: str) -> str:
    return _COLUMN_LABELS.get(col, col.replace("_", " ").capitalize())


def _auto_width(ws, columns: list[str], rows: list[dict]):
    """Устанавливает ширину колонок по содержимому (с ограничениями)."""
    MIN_WIDTH = 10
    MAX_WIDTH = 60

    for col_idx, col_name in enumerate(columns, 1):
        # ширина заголовка
        header_len = len(_humanize_column(col_name))
        max_len = header_len

        # ширина данных (смотрим первые 50 строк)
        for row_data in rows[:50]:
            val = _format_value(row_data.get(col_name, ""))
            if isinstance(val, str):
                # берём длину самой длинной строки (если многострочная)
                for line in str(val).split("\n"):
                    max_len = max(max_len, len(line))
            else:
                max_len = max(max_len, len(str(val)))

        width = min(max(max_len + 2, MIN_WIDTH), MAX_WIDTH)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


# ═════════════════════════════════════════════════════════════════════════════
# HTML (без изменений)
# ═════════════════════════════════════════════════════════════════════════════

_CHECK_TITLES = {
    "empty_pages": "Пустые страницы",
    "seo": "SEO-проблемы",
    "broken_links": "Битые ссылки",
    "images": "Изображения",
    "redirects": "Редиректы",
    "duplicates": "Дубликаты контента",
    "placeholders": "Заглушки и placeholder-тексты",
}


def save_html(results: dict[str, list[dict]],
              summaries: dict[str, str],
              *,
              base_url: str = "",
              filepath: str = "audit_report.html"):
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    total_issues = sum(len(rows) for rows in results.values())

    sections_html = ""
    toc_html = ""

    for check_name, rows in results.items():
        title = _CHECK_TITLES.get(check_name, check_name)
        count = len(rows)
        anchor = f"section-{check_name}"

        severity_class = "ok" if count == 0 else "issues"
        toc_html += (
            f'<li class="{severity_class}">'
            f'<a href="#{anchor}">{_esc(title)}</a>'
            f' <span class="badge">{count}</span>'
            f'</li>\n'
        )

        summary_text = summaries.get(check_name, "")
        sections_html += f'<section id="{anchor}">\n'
        sections_html += f'<h2>{_esc(title)} <span class="badge">{count}</span></h2>\n'

        if summary_text:
            sections_html += f'<pre class="summary">{_esc(summary_text)}</pre>\n'

        if rows:
            sections_html += _rows_to_html_table(rows)
        else:
            sections_html += '<p class="ok-msg">Проблем не найдено.</p>\n'

        sections_html += '</section>\n'

    page = _HTML_TEMPLATE.format(
        base_url=_esc(base_url),
        date=now,
        total_issues=total_issues,
        toc=toc_html,
        sections=sections_html,
    )

    with open(filepath, "w", encoding="utf-8") as f:
        f.write(page)

    print(f"  [report] HTML-отчёт сохранён: {filepath}")


def _rows_to_html_table(rows: list[dict]) -> str:
    if not rows:
        return ""

    columns: list[str] = []
    seen: set[str] = set()
    for row in rows:
        for key in row:
            if key not in seen and key != "check":
                columns.append(key)
                seen.add(key)

    lines = ['<div class="table-wrap"><table>\n<thead><tr>']
    for col in columns:
        lines.append(f"<th>{_esc(col)}</th>")
    lines.append("</tr></thead>\n<tbody>")

    for row in rows:
        lines.append("<tr>")
        for col in columns:
            val = row.get(col, "")
            cell = _format_html_cell(col, val)
            lines.append(f"<td>{cell}</td>")
        lines.append("</tr>")

    lines.append("</tbody>\n</table></div>")
    return "\n".join(lines)


def _format_html_cell(col: str, val) -> str:
    if val is None:
        return '<span class="empty">—</span>'
    if isinstance(val, bool):
        return "да" if val else ""
    if isinstance(val, list):
        if not val:
            return ""
        items = []
        for item in val:
            if isinstance(item, dict):
                items.append(_esc(json.dumps(item, ensure_ascii=False, default=str)))
            else:
                items.append(_esc(str(item)))
        return "<br>".join(items)
    if isinstance(val, dict):
        return _esc(json.dumps(val, ensure_ascii=False, default=str))

    s = str(val)
    if s.startswith("http://") or s.startswith("https://"):
        return f'<a href="{_esc(s)}" target="_blank" rel="noopener">{_esc(s)}</a>'
    return _esc(s)


def _esc(text: str) -> str:
    return html_mod.escape(str(text))


_HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Аудит сайта — {base_url}</title>
<style>
  :root {{
    --bg: #f8f9fa; --card: #ffffff; --border: #dee2e6;
    --accent: #0d6efd; --danger: #dc3545; --success: #198754;
    --text: #212529; --muted: #6c757d;
  }}
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{
    font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
    background: var(--bg); color: var(--text); line-height: 1.6;
    padding: 2rem; max-width: 1400px; margin: 0 auto;
  }}
  h1 {{ margin-bottom: 0.25rem; }}
  .meta {{ color: var(--muted); margin-bottom: 2rem; }}
  .stats {{ display: flex; gap: 1.5rem; margin-bottom: 2rem; flex-wrap: wrap; }}
  .stat-card {{
    background: var(--card); border: 1px solid var(--border);
    border-radius: 8px; padding: 1rem 1.5rem;
  }}
  .stat-card .num {{ font-size: 2rem; font-weight: 700; }}
  .stat-card .label {{ color: var(--muted); font-size: 0.9rem; }}
  nav ul {{ list-style: none; display: flex; flex-wrap: wrap; gap: 0.5rem; margin-bottom: 2rem; }}
  nav li {{
    background: var(--card); border: 1px solid var(--border);
    border-radius: 6px; padding: 0.4rem 0.8rem; font-size: 0.9rem;
  }}
  nav li.issues {{ border-left: 3px solid var(--danger); }}
  nav li.ok {{ border-left: 3px solid var(--success); }}
  nav a {{ text-decoration: none; color: var(--text); }}
  nav a:hover {{ color: var(--accent); }}
  .badge {{
    display: inline-block; background: var(--danger); color: #fff;
    border-radius: 10px; padding: 0.1rem 0.5rem; font-size: 0.8rem;
    vertical-align: middle;
  }}
  section {{
    background: var(--card); border: 1px solid var(--border);
    border-radius: 8px; padding: 1.5rem; margin-bottom: 1.5rem;
  }}
  section h2 {{ margin-bottom: 1rem; font-size: 1.3rem; }}
  .summary {{
    background: var(--bg); border: 1px solid var(--border);
    border-radius: 4px; padding: 1rem; margin-bottom: 1rem;
    font-size: 0.85rem; white-space: pre-wrap; overflow-x: auto;
  }}
  .ok-msg {{ color: var(--success); font-weight: 500; }}
  .table-wrap {{ overflow-x: auto; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 0.85rem; }}
  th, td {{
    border: 1px solid var(--border); padding: 0.5rem 0.6rem;
    text-align: left; vertical-align: top;
  }}
  th {{ background: var(--bg); position: sticky; top: 0; white-space: nowrap; }}
  tr:hover {{ background: #f1f3f5; }}
  a {{ color: var(--accent); word-break: break-all; }}
  .empty {{ color: var(--muted); }}
</style>
</head>
<body>
<h1>Аудит сайта</h1>
<p class="meta">{base_url} &middot; {date}</p>
<div class="stats">
  <div class="stat-card">
    <div class="num">{total_issues}</div>
    <div class="label">проблем найдено</div>
  </div>
</div>
<nav><ul>{toc}</ul></nav>
{sections}
</body>
</html>"""


# ═════════════════════════════════════════════════════════════════════════════
# Утилита: сохранить оба формата
# ═════════════════════════════════════════════════════════════════════════════

def save_all(results: dict[str, list[dict]],
             summaries: dict[str, str],
             *,
             base_url: str = "",
             output_dir: str = ".",
             excel_name: str = "audit_report.xlsx",
             html_name: str = "audit_report.html"):
    """Сохраняет Excel и HTML в указанную директорию."""
    os.makedirs(output_dir, exist_ok=True)
    excel_path = str(Path(output_dir) / excel_name)
    html_path = str(Path(output_dir) / html_name)

    save_excel(results, summaries, base_url=base_url, filepath=excel_path)
    save_html(results, summaries, base_url=base_url, filepath=html_path)
